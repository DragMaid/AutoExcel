#Leave this at the top in order to work
from kivy.config import Config
Config.set('graphics', 'resizable', False)
from kivymd.app import MDApp
from kivy.core.window import Window
Window.size = (950, 700)
from kivymd.uix.label import MDLabel
from kivy.uix.screenmanager import Screen, ScreenManager
from kivy.lang import Builder
from tkinter import filedialog
from tkinter import Tk
from kivy.properties import ObjectProperty, StringProperty, ListProperty, NumericProperty, BooleanProperty
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from openpyxl import load_workbook
from kivymd.uix.menu import MDDropdownMenu
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.graphics import Color
from kivymd.uix.toolbar import MDToolbar
from threading import Thread

ALPHABET = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]

class NavControl(ScreenManager):
    pass

class SelectScreen(Screen):
    pass

class LoadingScreen(Screen):
    pass

class ErrorScreen(Screen):
    pass

#Cell square Widget
class CustomSheetItems(Label):
    bgColor = ListProperty([])
    def __init__(self, bgColor, **kwargs):
        super().__init__(**kwargs)
        self.bgColor = bgColor

#Uses Kivymd Datatable to display xlsx files
class Sheet(GridLayout):
    #Color declaration
    gray = [213/255, 211/255, 200/255, 1]
    white = [1,1,1,1]

    #Board configuration values
    RowReq = 100
    defaultWidth = 150
    defaultNumWidth = 50
    defaultHeight = 50
    
    #Board widgets reference
    dataBoard = ListProperty()
    dummy = ListProperty()

    #General Excel file info
    filepath = StringProperty()
    filename = StringProperty()
    workbook = ""
    worksheet = ""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.setupGrid()

    #Create board cells before changing value
    def setupGrid(self):
        self.AddToSheet(text="",bgColor=self.gray, width=self.defaultNumWidth)
        for x in ALPHABET:
            self.AddToSheet(text=x,bgColor=self.gray)

        #Append cells' reference
        for y in range(1,self.RowReq+1):
            self.AddToSheet(text=y,bgColor=self.gray,width=self.defaultNumWidth)
            for st in range(26):
                self.AddToSheet(text="",bgColor=self.white)

        for items in self.dummy:
            self.add_widget(items) 
        self.dataBoard.append(self.dummy)

    def AddToSheet(self, text, bgColor, width=150, color="#484747", *args, **kwargs):
        self.dummy.append(CustomSheetItems(text=str(text),bgColor=bgColor,font_size=20,color=color,size_hint_x=None,size_hint_y=None,height=self.defaultHeight,width=width))
    
    #Set all cells's text to default
    def resetSheet(self):
        for x in range(1, self.RowReq+1):
            for y in range(1, 26):
                self.dataBoard[0][x*27+y].text = ""
        self.readData()

    def getMaxLen(self):
        maxRow = len(self.worksheet[1])
        maxCol = len(self.worksheet['A'])
        return maxRow, maxCol

    #Setting configurations when receive valid file
    def InitData(self, filepath, filename):
        self.filepath = filepath
        self.filename = filename
        self.reload_workbook()
        self.readData()


    #Assign cell's text to file data
    def readData(self):
        self.reload_workbook()
        maxRow, maxCol = self.getMaxLen()
        for x in range(1, maxCol+1):
            for y in range(1, len(self.worksheet[x])+1):
                var = str((self.worksheet.cell(row=x, column=y).value))
                if var == "None":
                    self.dataBoard[0][x*27+y].text = ""
                elif len(var) > 9:
                    self.databoard[0][x*27+y].text = var[0:9] + "..."
                else:
                    self.dataBoard[0][x*27+y].text = var

    #Get changes and apply
    def reload_workbook(self):
        self.workbook = load_workbook(self.filepath)
        self.worksheet = self.workbook.active

    def Auto_Add(self, mode):
        self.reload_workbook()
        maxRow, maxCol = self.getMaxLen()
        if mode == "Row":
            for y in range(2, maxRow+1):
                maxTry = 3
                total = 0
                for x in range(1, maxCol+1):
                    local = self.worksheet.cell(row=x, column=y).value
                    if type(local) == int or type(local) == float:
                        total += local
                    else:
                        maxTry -= 1
                    if maxTry == 0:
                        break
                self.worksheet.cell(row=maxCol+1, column=y, value=total)
                self.dataBoard[0][(maxCol+1)*27+y].text = str(total)
        else:
            for x in range(2, maxCol+1):
                maxTry = 3
                total = 0
                for y in range(1, maxRow+1):
                    local = self.worksheet.cell(row=x, column=y).value
                    if type(local) == int or type(local) == float:
                        total += local
                    else:
                        maxTry -= 1
                    if maxTry == 0:
                        break
                self.worksheet.cell(row=x, column=maxRow+1, value=total)
                self.dataBoard[0][x*27+maxRow+1].text = str(total)

    def saveFile(self):
        self.workbook.save(self.filepath)

class MainScreen(Screen):
    icon = StringProperty("align-vertical-top")
    mode = StringProperty("Row")

    def __init__(self, **kwargs):
        super(MainScreen, self).__init__(**kwargs)
        self.menuItems = [
            {"text":"Auto","viewclass":"OneLineListItem"},
            {"text":"Custom","viewclass":"OneLineListItem"},
        ]

    def showOptions(self, target, *args):
        self.menu = MDDropdownMenu(
            max_height=100,
            caller=target,
            items=self.menuItems,
            width_mult=2.5,
        )
        self.menu.open()

    def changeMode(self):
        if self.icon != "align-horizontal-left":
            self.icon = "align-horizontal-left"
            self.mode = "Col"
        else:
            self.icon = "align-vertical-top"
            self.mode = "Row"

#Receive user's file request
class DropSquare(FloatLayout):
    filepath = StringProperty()
    filename = StringProperty()
    def __init__(self, **kwargs):
        super(DropSquare, self).__init__(**kwargs)

        # get app instance to add function from widget
        app = MDApp.get_running_app()

        # add function to the list
        app.drops.append(self.on_dropfile)

    def FileManagerOpen(self, *args):
        #Using tkinter filemanager as an alternative to kivy's
        root = Tk()
        root.withdraw()
        filepath = filedialog.askopenfilename(initialdir = "/",
                                              title = "Select a File",
                                              filetypes = (
                                                [
                                                  "Excel extensions",
                                                  ".xlsx .xlsm .xltx .xltm",
                                                ],
                                                (
                                                  "all files",
                                                  "*.*"
                                                )
                                              )
                                                    )
        root.destroy()
        self.switchScreen('LoadingScreen')
        thread = Thread(target=self.ProccessFile, args=(str(filepath),), daemon=True)
        thread.start()

    def ProccessFile(self, filepath):
        try:
            load_workbook(filepath)
            self.filepath = filepath
            self.getFilename()
            self.ref.manager.ids.MainScreen.ids.Sheet.InitData(filepath=self.filepath, filename=self.filename)
            self.switchScreen("MainScreen")
        except:
            self.switchScreen("ErrorScreen")

    def on_dropfile(self, widget, filename):
        if self.collide_point(*Window.mouse_pos):
            # on_dropfile's filename is bytes (py3)
            decodedPath = filename.decode()
            self.switchScreen("LoadingScreen")
            thread = Thread(target=self.ProccessFile, args=(decodedPath,), daemon=True)
            thread.start()

    def getFilename(self):
        if self.filepath.split('\\')[-1] == self.filepath:  # "\\" is for Window only = '\'
            self.filename = self.filepath.split('/')[-1]
        else:
            self.filename = self.filepath.split('\\')[-1]

    def switchScreen(self, Wscreen, *args):
        app = MDApp.get_running_app()
        app.root.current = str(Wscreen)

class MainApp(MDApp):
    def build(self):
        self.drops = []
        # bind handling function to 'on_dropfile'
        Window.bind(on_dropfile=self.handledrops)
        self.root = Builder.load_file("display.kv")
        return self.root

    def handledrops(self, *args):
        # this will execute each function from list with arguments from
        # Window.on_dropfile
        #
        # make sure `Window.on_dropfile` works on your system first,
        # otherwise the example won't work at all
        for func in self.drops:
            func(*args)

if __name__ == "__main__":
    MainApp().run()